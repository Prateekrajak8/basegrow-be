import csv
import io
import os
import time
from typing import Any, Dict, List

import requests
from django.conf import settings
from django.db import connection
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, parser_classes, authentication_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.authentication import SessionAuthentication

from .services import (
    build_date_filter,
    cache,
    calc_campaign_ttl,
    get_cache_key,
    ongage_get,
    ongage_post,
    serialize,
)


def _dictfetchall(cursor) -> List[Dict[str, Any]]:
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


class CsrfExemptSessionAuthentication(SessionAuthentication):
    def enforce_csrf(self, request):
        return


@api_view(["GET", "POST"])
def mailcamp_clusters_view(request):
    if request.method == "GET":
        try:
            account_id = int(request.query_params.get("accountId", ""))
        except Exception:
            return Response({"error": "Invalid accountId"}, status=400)

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    'SELECT id, "accountId", name, domains, "createdAt", "updatedAt" FROM "Cluster" WHERE "accountId" = %s ORDER BY name ASC',
                    [account_id],
                )
                clusters = _dictfetchall(cursor)
            return Response({"clusters": serialize(clusters)}, status=200)
        except Exception:
            return Response({"error": "Failed to fetch clusters"}, status=500)

    try:
        cluster_id = request.data.get("id")
        if not cluster_id:
            return Response({"error": "Cluster ID is required"}, status=400)
        return Response({"cluster": ""}, status=200)
    except Exception:
        return Response({"error": "Failed to update cluster"}, status=500)


@api_view(["GET"])
def fetch_campaign_data_view(request):
    try:
        account_id_param = request.query_params.get("accountId")
        params: List[Any] = []
        where = ""
        if account_id_param:
            where = 'WHERE c."accountId" = %s'
            params.append(int(account_id_param))

        with connection.cursor() as cursor:
            cursor.execute(
                f'''
                SELECT
                    c.id,
                    c."statId",
                    c.name,
                    c.sent,
                    c."hardBounce",
                    c."softBounce",
                    c."uniqueOpeners",
                    c."totalClicks",
                    c."uniqueClickers",
                    c.unsubscribes,
                    c.complaints,
                    c."createdAt",
                    c."accountId",
                    a.name AS account_name,
                    a."apiKey" AS account_api_key
                FROM "Campaign" c
                LEFT JOIN "Account" a ON a.id = c."accountId"
                {where}
                ORDER BY c."createdAt" DESC
                ''',
                params,
            )
            rows = _dictfetchall(cursor)

        campaigns = []
        for row in rows:
            account = None
            if row.get("accountId") is not None:
                account = {
                    "id": row.get("accountId"),
                    "name": row.get("account_name"),
                    "apiKey": row.get("account_api_key"),
                }
            row.pop("account_name", None)
            row.pop("account_api_key", None)
            row["account"] = account
            campaigns.append(row)

        return Response(serialize(campaigns), status=200)
    except Exception:
        return Response({"error": "Failed to fetch campaigns"}, status=500)


@api_view(["POST"])
def get_cluster_data_view(request):
    names = request.data.get("names")
    stat_ids = request.data.get("statIds")
    if not names and not stat_ids:
        return Response({"error": "Either names or statIds required"}, status=400)
    return Response("", status=200)


@api_view(["POST"])
@authentication_classes([CsrfExemptSessionAuthentication])
def domain_data_by_campaign_view(request):
    try:
        campaign_id = int(request.data.get("campaignId"))
    except Exception:
        return Response({"success": False, "error": "campaignId is required"}, status=400)

    try:
        cache_key = get_cache_key("domain-data", {"campaignId": campaign_id})
        cached = cache.get(cache_key)
        if cached:
            return Response(cached, status=200)

        with connection.cursor() as cursor:
            cursor.execute(
                '''
                SELECT
                    dis.domain,
                    dis.sent,
                    dis.open,
                    dis.click,
                    dis."hardBounce",
                    dis."softBounce",
                    dis."uniqueOpen",
                    dis."uniqueClick"
                FROM "DomainInteractionSummary" dis
                WHERE dis."campaignId" = %s
                ORDER BY dis.domain ASC
                ''',
                [campaign_id],
            )
            rows = _dictfetchall(cursor)

        formatted = []
        for row in rows:
            sent = int(row.get("sent") or 0)
            unique_open = int(row.get("uniqueOpen") or 0)
            unique_click = int(row.get("uniqueClick") or 0)
            click = int(row.get("click") or 0)
            formatted.append(
                {
                    "domainName": row.get("domain") or "Unknown",
                    "sent": sent,
                    "uniqueOpen": unique_open,
                    "totalClicks": click,
                    "softBounce": int(row.get("softBounce") or 0),
                    "hardBounce": int(row.get("hardBounce") or 0),
                    "uniqueClicks": unique_click,
                    "uniqueOpenRate": round((unique_open / sent) * 100, 2) if sent else 0,
                    "uniqueClicksRate": round((unique_click / sent) * 100, 2) if sent else 0,
                    "toalClickRate": round((click / sent) * 100, 2) if sent else 0,
                }
            )

        with connection.cursor() as cursor:
            cursor.execute('SELECT name FROM "Campaign" WHERE id = %s LIMIT 1', [campaign_id])
            row = cursor.fetchone()
            campaign_name = row[0] if row else ""

        payload = serialize({"success": True, "data": formatted})
        ttl = calc_campaign_ttl(campaign_name)
        cache.set(cache_key, payload, ttl)
        return Response(payload, status=200)
    except Exception as exc:
        return Response({"success": False, "error": str(exc)}, status=500)


@api_view(["POST"])
@authentication_classes([CsrfExemptSessionAuthentication])
def cluster_data_by_campaign_view(request):
    try:
        campaign_id = int(request.data.get("campaignId"))
    except Exception:
        return Response({"success": False, "error": "campaignId is required"}, status=400)

    try:
        cache_key = get_cache_key("cluster-data", {"campaignId": campaign_id})
        cached = cache.get(cache_key)
        if cached:
            return Response(cached, status=200)

        with connection.cursor() as cursor:
            cursor.execute(
                '''
                SELECT
                    cis."clusterId",
                    c.name AS "clusterName",
                    cis.sent,
                    cis.open,
                    cis.click,
                    cis."hardBounce",
                    cis."softBounce",
                    cis."uniqueOpen",
                    cis."uniqueClick"
                FROM "ClusterInteractionSummary" cis
                LEFT JOIN "Cluster" c ON cis."clusterId" = c.id
                WHERE cis."campaignId" = %s
                ORDER BY cis."clusterId" ASC
                ''',
                [campaign_id],
            )
            rows = _dictfetchall(cursor)

        formatted = []
        for row in rows:
            sent = int(row.get("sent") or 0)
            unique_open = int(row.get("uniqueOpen") or 0)
            unique_click = int(row.get("uniqueClick") or 0)
            click = int(row.get("click") or 0)
            formatted.append(
                {
                    "clusterId": int(row.get("clusterId") or 0),
                    "clusterName": row.get("clusterName") or "Unknown",
                    "sent": sent,
                    "uniqueOpen": unique_open,
                    "uniqueClicks": unique_click,
                    "totalClicks": click,
                    "softBounce": int(row.get("softBounce") or 0),
                    "hardBounce": int(row.get("hardBounce") or 0),
                    "uniqueOpenRate": round((unique_open / sent) * 100, 2) if sent else 0,
                    "uniqueClicksRate": round((unique_click / sent) * 100, 2) if sent else 0,
                    "toalClickRate": round((click / sent) * 100, 2) if sent else 0,
                }
            )

        with connection.cursor() as cursor:
            cursor.execute('SELECT name FROM "Campaign" WHERE id = %s LIMIT 1', [campaign_id])
            row = cursor.fetchone()
            campaign_name = row[0] if row else ""

        payload = serialize({"success": True, "data": formatted})
        ttl = calc_campaign_ttl(campaign_name)
        cache.set(cache_key, payload, ttl)
        return Response(payload, status=200)
    except Exception as exc:
        return Response({"success": False, "error": str(exc)}, status=500)


def _ongage_base_payload(request_data, from_key, group, select, extra_filters=None, list_id="all"):
    time_filter = request_data.get("time_filter", "all_time")
    start_date = request_data.get("start_date")
    end_date = request_data.get("end_date")

    date_filter = build_date_filter(time_filter, start_date, end_date)
    filters = [
        ["is_test_campaign", "=", 0],
        ["email_message_type", "=", "email_message"],
        *date_filter,
    ]

    if extra_filters:
        filters.extend(extra_filters)

    return {
        "filter": filters,
        "from": from_key,
        "group": group,
        "list_id": request_data.get("list_id", list_id),
        "order": [["MAX(`delivery_timestamp`)", "desc"]],
        "select": select,
        "time_zone": "UTC",
    }


@api_view(["POST"])
def ongage_matrix_view(request):
    try:
        payload = _ongage_base_payload(
            request.data,
            from_key="mailing",
            group=["list_id", "stats_date"],
            select=[
                "list_id",
                ["MAX(`stats_date`)", "stats_date"],
                ["MAX(`delivery_timestamp`)", "delivery_timestamp"],
                "sum(`targeted`)",
                "sum(`sent`)",
                "sum(`success`)",
                "sum(`failed`)",
                "sum(`hard_bounces`)",
                "sum(`soft_bounces`)",
                "sum(`opens`)",
                "sum(`unique_opens`)",
                "sum(`clicks`)",
                "sum(`unique_clicks`)",
                "sum(`unsubscribes`)",
                "sum(`complaints`)",
                "esp_name_title",
                "isp_name_or_others",
                "sum(`post_back_clicks`)",
                "ctr",
                "uctr",
            ],
        )
        data = ongage_post("https://api.ongage.net/api/reports/query", payload)
        return Response(data, status=200)
    except Exception as exc:
        return Response({"message": "Something went wrong", "error": str(exc)}, status=500)


@api_view(["POST"])
def ongage_events_view(request):
    try:
        domain = request.data.get("domain")
        extra = []
        if domain and domain != "All":
            extra.append(["isp_name_or_others", "=", domain])

        payload = _ongage_base_payload(
            request.data,
            from_key="event",
            group=["event_name"],
            select=[
                "list_id",
                "segment_name",
                "event_name",
                "event_id",
                ["MAX(`stats_date`)", "stats_date"],
                ["MAX(`delivery_timestamp`)", "delivery_timestamp"],
                "sum(`targeted`)",
                "sum(`sent`)",
                "sum(`success`)",
                "sum(`failed`)",
                "sum(`hard_bounces`)",
                "sum(`soft_bounces`)",
                "sum(`opens`)",
                "sum(`unique_opens`)",
                "sum(`clicks`)",
                "sum(`unique_clicks`)",
                "sum(`unsubscribes`)",
                "sum(`complaints`)",
                "esp_name_title",
                "isp_name_or_others",
                "sum(`post_back_clicks`)",
                "ctr",
                "uctr",
            ],
            extra_filters=extra,
        )
        data = ongage_post("https://api.ongage.net/api/reports/query", payload)
        return Response(data, status=200)
    except Exception as exc:
        return Response({"message": "Something went wrong", "error": str(exc)}, status=500)


@api_view(["POST"])
def ongage_event_date_data_view(request):
    try:
        event_name = request.data.get("event_name")
        payload = _ongage_base_payload(
            request.data,
            from_key="mailing",
            group=["stats_date"],
            select=[
                "list_id",
                "segment_name",
                "segment_id",
                "event_name",
                "event_id",
                ["MAX(stats_date)", "stats_date"],
                ["MAX(delivery_timestamp)", "delivery_timestamp"],
                "sum(targeted)",
                "sum(sent)",
                "sum(success)",
                "sum(failed)",
                "sum(hard_bounces)",
                "sum(soft_bounces)",
                "sum(opens)",
                "sum(unique_opens)",
                "sum(clicks)",
                "sum(unique_clicks)",
                "sum(unsubscribes)",
                "sum(complaints)",
                "esp_name_title",
                "isp_name_or_others",
                "sum(post_back_clicks)",
                "ctr",
                "uctr",
            ],
            extra_filters=[["event_name", "=", event_name]],
        )
        data = ongage_post("https://api.ongage.net/api/reports/query", payload)
        return Response(data, status=200)
    except Exception as exc:
        return Response({"message": "Something went wrong", "error": str(exc)}, status=500)


@api_view(["POST"])
def ongage_events_name_data_view(request):
    try:
        segment_name = request.data.get("segment_name")
        payload = _ongage_base_payload(
            request.data,
            from_key="mailing",
            group=["stats_date"],
            select=[
                "list_id",
                "segment_name",
                ["MAX(stats_date)", "stats_date"],
                ["MAX(delivery_timestamp)", "delivery_timestamp"],
                "sum(targeted)",
                "sum(sent)",
                "sum(success)",
                "sum(failed)",
                "sum(hard_bounces)",
                "sum(soft_bounces)",
                "sum(opens)",
                "sum(unique_opens)",
                "sum(clicks)",
                "sum(unique_clicks)",
                "sum(unsubscribes)",
                "sum(complaints)",
                "esp_name_title",
                "isp_name_or_others",
                "sum(post_back_clicks)",
                "ctr",
                "uctr",
            ],
            extra_filters=[["segment_name", "=", segment_name]],
            list_id="all",
        )
        data = ongage_post("https://api.ongage.net/api/reports/query", payload)
        return Response(data, status=200)
    except Exception as exc:
        return Response({"message": "Something went wrong", "error": str(exc)}, status=500)


@api_view(["POST"])
def ongage_segments_view(request):
    try:
        event_name = request.data.get("event_name")
        smtp_id = request.data.get("smtp_id")
        extra = [["event_name", "=", event_name]]
        if smtp_id:
            extra.append(["esp_connection_id", "=", int(smtp_id)])

        payload = _ongage_base_payload(
            request.data,
            from_key="mailing",
            group=["segment_name"],
            select=[
                "list_id",
                "segment_name",
                "segment_id",
                "esp_connection_title",
                "esp_connection_id",
                ["MAX(`stats_date`)", "stats_date"],
                ["MAX(`delivery_timestamp`)", "delivery_timestamp"],
                "sum(`targeted`)",
                "sum(`sent`)",
                "sum(`success`)",
                "sum(`failed`)",
                "sum(`hard_bounces`)",
                "sum(`soft_bounces`)",
                "sum(`opens`)",
                "sum(`unique_opens`)",
                "sum(`clicks`)",
                "sum(`unique_clicks`)",
                "sum(`unsubscribes`)",
                "sum(`complaints`)",
                "esp_name_title",
                "isp_name_or_others",
                "sum(`post_back_clicks`)",
                "ctr",
                "uctr",
            ],
            extra_filters=extra,
        )
        data = ongage_post("https://api.ongage.net/api/reports/query", payload)
        return Response(data, status=200)
    except Exception as exc:
        return Response({"message": "Something went wrong", "error": str(exc)}, status=500)


@api_view(["GET"])
def ongage_get_esp_connection_view(request):
    try:
        data = ongage_get("https://api.ongage.net/api/esp_connections/options")
        return Response(data, status=200)
    except Exception as exc:
        return Response({"message": "Something went wrong", "error": str(exc)}, status=500)


@api_view(["POST"])
def ongage_event_status_view(request):
    try:
        event_id = request.data.get("event_id")
        list_id = request.data.get("list_id")
        if not event_id or not list_id:
            return Response({"message": "Missing event_id or list_id in body"}, status=400)

        headers = {
            "x_account_code": settings.ONGAGE_ACCOUNT_CODE,
            "x_username": settings.ONGAGE_USERNAME,
            "x_password": settings.ONGAGE_PASSWORD,
        }
        resp = requests.get(
            f"https://api.ongage.net/api/events/{event_id}?list_id={list_id}",
            headers=headers,
            timeout=60,
        )

        if resp.status_code >= 400:
            try:
                error_data = resp.json()
            except Exception:
                error_data = resp.text
            return Response({"message": "Failed to fetch event", "error": error_data}, status=resp.status_code)

        payload = resp.json().get("payload", {})
        return Response(
            {
                "event_id": event_id,
                "time_to_send_config": payload.get("time_to_send_config"),
                "status": payload.get("status"),
                "status_desc": payload.get("status_desc"),
                "status_date": payload.get("status_date"),
            },
            status=200,
        )
    except Exception as exc:
        return Response({"message": "Server error", "error": str(exc)}, status=500)


@api_view(["POST"])
def ongage_segement_count_view(request):
    try:
        list_id = request.data.get("list_id")
        segment_id = request.data.get("segment_id")
        if not list_id or not segment_id:
            return Response({"error": "Missing list_id or segment_id"}, status=400)

        headers = {
            "x_account_code": settings.ONGAGE_ACCOUNT_CODE,
            "x_username": settings.ONGAGE_USERNAME,
            "x_password": settings.ONGAGE_PASSWORD,
        }

        start_resp = requests.post(
            f"https://api.ongage.net/{int(list_id)}/api/contact_counts",
            headers={"Content-Type": "application/json", **headers},
            json={"segment_id": int(segment_id)},
            timeout=60,
        )
        start_resp.raise_for_status()

        data = start_resp.json()
        count_id = data.get("payload", {}).get("id")
        if not count_id:
            return Response({"error": "Count ID not returned"}, status=500)

        retries = 0
        max_retries = 30
        while retries < max_retries:
            poll_resp = requests.get(
                f"https://api.ongage.net/api/contact_counts/{int(count_id)}",
                headers=headers,
                timeout=60,
            )
            poll_resp.raise_for_status()
            poll_data = poll_resp.json()
            if poll_data.get("metadata", {}).get("total") is not None:
                return Response(poll_data, status=200)
            time.sleep(2)
            retries += 1

        return Response({"error": "Polling timed out. metadata.total still null."}, status=504)
    except Exception as exc:
        return Response({"error": "Failed to fetch segment count", "detail": str(exc)}, status=500)


@api_view(["GET"])
def spreadsheet_url_view(request):
    return Response({"url": settings.GOOGLE_SPREADSHEET_URL}, status=200)


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def import_view(request):
    try:
        for_org = request.data.get("for_org")
        uploaded_by = request.data.get("uploaded_by")

        if not for_org or not uploaded_by:
            return Response({"error": "Missing required fields"}, status=400)

        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=400)

        ext = os.path.splitext(file_obj.name)[1].lower()
        parsed_rows = []

        if ext in (".xlsx", ".xls"):
            wb = load_workbook(file_obj, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h is not None else "" for h in rows[0]]
                for row in rows[1:]:
                    parsed_rows.append({headers[i]: row[i] for i in range(len(headers))})
        elif ext == ".csv":
            content = file_obj.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(content))
            parsed_rows = list(reader)
        else:
            return Response(
                {"error": "Invalid file format. Please upload .xlsx, .xls, or .csv file"},
                status=400,
            )

        import_result = {
            "for_org": for_org,
            "uploaded_by": uploaded_by,
            "file_type": ext.lstrip("."),
            "rows_count": len(parsed_rows),
            "sample_data": parsed_rows[:3],
        }

        return Response(
            {
                "success": True,
                "message": "File imported successfully",
                "result": serialize(import_result),
            },
            status=200,
        )
    except Exception as exc:
        return Response({"error": "Failed to import file", "details": str(exc)}, status=500)


@api_view(["GET"])
def country_view(request):
    return Response([], status=200)


@api_view(["GET"])
def esp_view(request):
    return Response([], status=200)


@api_view(["GET"])
def domain_view(request):
    return Response([], status=200)


@api_view(["GET"])
def isp_view(request):
    return Response({"data": []}, status=200)


@api_view(["POST"])
def add_isp_view(request):
    return Response({"message": "Not implemented yet"}, status=200)


@api_view(["GET"])
def users_get_users_view(request):
    return Response([], status=200)


@api_view(["POST"])
def users_onboard_view(request):
    return Response({"message": "Not implemented yet"}, status=201)


@api_view(["POST"])
def users_set_password_view(request):
    return Response({"message": "Not implemented yet"}, status=200)


@api_view(["GET", "POST"])
def click_stats_get_user_data_view(request):
    return Response([], status=200)


@api_view(["POST"])
def click_stats_import_view(request):
    return Response({"message": "Not implemented yet"}, status=201)


@api_view(["POST"])
def mailcamp_fetch_user_data_id_view(request):
    return Response([], status=200)


@api_view(["GET"])
def sheet_data_view(request):
    return Response({"data": []}, status=200)
